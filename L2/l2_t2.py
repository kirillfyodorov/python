import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import random
# объект
wb = Workbook()
# активный лист
ws = wb.active
def table(wb,ws):     # заполнение таблицы
    n = 20
    for i in range(1,n+1):
        for j in range(1,n+1):
            r = random.randint(1,30)
            ws.cell(column=j, row=i, value=r)
    #сохранение файла в текущую директорию
    wb.save("lab.xlsx")
    x = []
    y = []
    for i in range(10):
        x.append(random.randint(1,10))
        y.append(random.randint(1,10))
    # удаляем 10 элементов
    print(x)
    print(y)
    for i in range(len(x)):
        ws[get_column_letter(y[i]) + str(x[i])] = 0
    wb.save("lab.xlsx")

def winsorizing(wb, ws):
    for i in range(1,20):
        for j in range(1,20):
            if j==1 and ws[get_column_letter(j) + str(i)].value == 0: #для первого столбца используем ячейки правее
                r = ws[get_column_letter(j+1) + str(i)].value
                ws.cell(column=j, row=i, value=r)
                continue
            if ws[get_column_letter(j) + str(i)].value == 0: #для восстановления других столбцов используем ячейки левее
                r = ws[get_column_letter(j-1) + str(i)].value
                ws.cell(column=j, row=i, value=r)
        wb.save("labwinsorizing.xlsx")

def correlation(wb, ws):
    p_str = input("Введите номер ряда, который должен быть восстановлен:\n")
    p = int(p_str)
    v_str = input("Введите номер ряда, который коррелирует с восстанавливаемым:\n")
    v = int(v_str)
    for i in range(1,20):
        if ws[get_column_letter(i) + str(p)].value == 0 and i==1: #для воостановления ячейки из первого столбца используем столбец правее
            ws.cell(column=i, row=p, value=ws[get_column_letter(i+1) + str(p)].value*ws[get_column_letter(i) + str(v)].value/ws[get_column_letter(i+1) + str(v)].value)			
        elif ws[get_column_letter(i) + str(p)].value == 0: #для восстановления других столбцов используем ячейки левее
            ws.cell(column=i, row=p, value=ws[get_column_letter(i-1) + str(p)].value*ws[get_column_letter(i) + str(v)].value/ws[get_column_letter(i-1) + str(v)].value)	
    wb.save("labcorrelation.xlsx")

def linap(wb, ws):
	for i in range(1,20):
		sumx = 0 #составляющие для расчета коэффицентов
		sumx2 = 0
		sumy = 0
		sumxy = 0
		for j in range(1,20):
			sumx += j - 1
			sumx2 += (j - 1)**2
			sumy += ws[get_column_letter(i) + str(j)].value
			sumxy += ws[get_column_letter(i) + str(j)].value*(j-1)
		a = (20*sumxy - sumx*sumy)/(20*sumx2 - sumx**2) #расчет коэффицентов
		b = (sumy-a*sumx)/20
		for j in range(1,20):
			if ws[get_column_letter(i) + str(j)].value == 0:
				r = a*j + b #расчет значения восстанавливаемой ячейки
				ws.cell(column=i, row=j, value=r)
	wb.save("lablineap.xlsx")


table(wb, ws)
winsorizing(wb, ws)
#correlation(wb, ws)
#linap(wb, ws)

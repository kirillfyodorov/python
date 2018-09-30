import random
import matplotlib.pyplot as plt
from openpyxl import Workbook


print('Введите минимальные и максимальные значения x и y,\nчтобы разница между минимальным и максимальным значением было не меньше 20.')

minX = int(input('Введите минимальное значение X:\n'))
maxX = int(input('Введите максимальное значение X:\n'))
minY = int(input('Введите минимальное значение Y:\n'))
maxY = int(input('Введите минимальное значение Y:\n'))
x = []
y = []
n = 20


def myF(minV, maxV, masV):
    i = 0
    while i < n:
        log = 0
        t = random.randint(minV, maxV)
        for j in range(len(masV)):
            if t == masV[j]:
                log = 1
                break
        if log == 0:
            masV.append(t)
            i = i + 1
    return (masV)


x = myF(minX, maxX, x)
y = myF(minY, maxY, y)


def xl():
    wb = Workbook()
    ws = wb.active

    ws['A1'] = 'x'
    ws['A2'] = 'y'

    for col in range(2, 22):
        ws.cell(column=col, row=1, value=x[col-2])
        ws.cell(column=col, row=2, value=y[col-2])

    wb.save("Data.xlsx")


xl()


def sumOne(a):
    i = 0
    sum1 = 0
    while i < n-1:
        sum1 = sum1 + a[i]
        i = i + 1
    return (sum1)


def sumTwo(a, b):
    i = 0
    sum2 = 0
    while i < n-1:
        sum2 = sum2 + a[i]*b[i]
        i = i + 1
    return (sum2)


k = (n * sumTwo(x, y) - sumOne(x)*sumOne(y))/(n*sumTwo(x, x)-sumTwo(x, x))
c = (sumOne(y) - k * sumOne(x))/n


y_coor = []

for key in x:
    y_coor.append(k*key + c)

plt.figure()
plt.scatter(x, y)
plt.plot(x, y_coor)
plt.ylabel('Y')
plt.xlabel('X')
plt.show()
